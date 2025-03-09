#for the fuctions of the webscrapping 
import undetected_chromedriver as uc
from openpyxl import load_workbook
from config import profile_path, chromedriver_path, xlsx_path, PHONEs_COLs
import bs4


def open_chrome_with_profile():
    # Create a new Chrome session with the Chrome profile

    # options = Options()
    options = uc.ChromeOptions()
    options.add_argument("--user-data-dir=" + profile_path)

    # Create a new instance of the Chrome driver with the specified options
    # driver = webdriver.Chrome(executable_path=chromedriver_path, chrome_options=options)
    driver = uc.Chrome(driver_executable_path=chromedriver_path, options=options)
    return driver


def open_xlsx_file():
    # Open Excel file and return the workbook and worksheet

    wb = load_workbook(filename=xlsx_path)
    ws = wb.active
    return wb, ws


def write_phones_to_xlsx_file(wb, ws, phones, row):
    # Write phones to Excel file

    for i in range(len(phones)):
        ws[PHONEs_COLs[i] + str(row)].value = phones[i]

    wb.save(xlsx_path)


def extract_phones_from_page(page_source):
    # Extract phones from the page source and return them as a list of strings

    phones = []
    try:
        # find all phones
        soup = bs4.BeautifulSoup(page_source, "html.parser")
        # find all a tags with title containing "Call"
        a_tags = soup.find_all("a", title=lambda x: x and "Search people with phone number" in x)
        for a_tag in a_tags:
            phone = a_tag.text.strip()
            phones.append(phone)

        return phones

    except Exception as e:
        print(str(e))
        return phones